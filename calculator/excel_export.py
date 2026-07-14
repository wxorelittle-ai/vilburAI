"""
Экспорт расчёта себестоимости в Excel — PRO-фича (раздел Модуль C ТЗ:
«Экспорт расчёта в Excel (PRO-тариф)»).
"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

NAVY = 'FF1E3A5F'
ORANGE = 'FFFF6B35'


def _header_cell(ws, cell, text):
    ws[cell] = text
    ws[cell].font = Font(bold=True, color='FFFFFFFF', size=11)
    ws[cell].fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type='solid')
    ws[cell].alignment = Alignment(horizontal='left', vertical='center')


def generate_raschet_xlsx(raschet) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Расчёт себестоимости'

    ws.merge_cells('A1:B1')
    ws['A1'] = f'Расчёт себестоимости от {raschet.data:%d.%m.%Y}'
    ws['A1'].font = Font(bold=True, size=14, color=NAVY)
    ws.row_dimensions[1].height = 24

    row = 3
    _header_cell(ws, f'A{row}', 'Вводные данные')
    ws.merge_cells(f'A{row}:B{row}')
    row += 1

    vvodnye = [
        ('Площадь помещения', f'{raschet.ploshad} м²'),
        ('Тип ремонта', raschet.get_tip_remonta_display()),
        ('Количество рабочих', raschet.kolvo_rabochih),
        ('Срок выполнения', f'{raschet.dni} дн.'),
        ('Оплата 1 рабочему в день', f'{raschet.stavka_v_den} ₽'),
        ('Аренда инструмента', f'{raschet.arenda} ₽'),
        ('Доставка материалов', f'{raschet.dostavka} ₽'),
        ('Расходники', f'{raschet.rashodniki} ₽'),
        ('Налоговый режим', raschet.get_nalog_display()),
    ]
    for label, value in vvodnye:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        row += 1

    row += 1
    _header_cell(ws, f'A{row}', 'Результаты расчёта')
    ws.merge_cells(f'A{row}:B{row}')
    row += 1

    rezultaty = [
        ('Трудозатраты', f'{raschet.trudozatraty} ₽'),
        ('Постоянные расходы', f'{raschet.postoyannye_raskhody} ₽'),
        ('Себестоимость (итого)', f'{raschet.sebestoimost_obshaya} ₽'),
        ('Себестоимость за м²', f'{raschet.sebestoimost_m2} ₽/м²'),
        ('Рекомендуемая цена +30%', f'{raschet.cena_30} ₽ ({raschet.cena_30_m2} ₽/м²)'),
        ('Рекомендуемая цена +50%', f'{raschet.cena_50} ₽ ({raschet.cena_50_m2} ₽/м²)'),
        ('Точка безубыточности', f'{raschet.tochka_bezubytochnosti_dney} дн.' if raschet.tochka_bezubytochnosti_dney else '—'),
        ('Ориентировочная цена по рынку', f'{raschet.rynochnaya_cena_m2} ₽/м²'),
        ('Отклонение от рынка', f'{raschet.otklonenie_ot_rynka_procent}%' if raschet.otklonenie_ot_rynka_procent is not None else '—'),
    ]
    for label, value in rezultaty:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        if label in ('Себестоимость (итого)', 'Себестоимость за м²'):
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].font = Font(bold=True, color='FFB00020')
        row += 1

    ws.column_dimensions[get_column_letter(1)].width = 32
    ws.column_dimensions[get_column_letter(2)].width = 40

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
