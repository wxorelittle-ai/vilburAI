"""Экспорт чеков ФНС в Excel для бухгалтера (раздел «Модуль D» ТЗ)."""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def cheki_v_excel(brigada, cheki) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Чеки ФНС'

    ws['A1'] = f'Чеки ФНС — {brigada.nazvanie}'
    ws['A1'].font = Font(size=13, bold=True)
    ws.append([])

    headers = ['Дата', 'Сумма, ₽', 'Назначение', 'Заказчик (тел.)', 'Статус', 'ID в ФНС', 'Демо']
    ws.append(headers)
    head_fill = PatternFill('solid', fgColor='262421')
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = head_fill

    itogo = 0
    for c in cheki:
        ws.append([
            c.data.strftime('%d.%m.%Y'), float(c.summa), c.naznachenie,
            c.telefon_zakazchika, c.get_status_display(), c.fns_id, 'да' if c.demo_rezhim else 'нет',
        ])
        itogo += float(c.summa)

    ws.append([])
    ws.append(['ИТОГО', itogo])
    ws[f'A{ws.max_row}'].font = Font(bold=True)
    ws[f'B{ws.max_row}'].font = Font(bold=True)

    widths = [12, 14, 40, 20, 20, 22, 8]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
